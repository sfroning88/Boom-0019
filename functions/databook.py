def generate_data_book(employees=None, years=None):
    """
    Generate data book xlsx file from processed employee data.
    Creates the format shown in the template with annual columns populated.
    
    Args:
        employees_data: Dictionary of employee payroll data by year
        
    Returns:
        BytesIO object containing the xlsx file or None if generation fails
    """

    if len(employees.keys()) == 0:
        print("ERROR: No employees have been processed yet")
        return None

    if len(years) == 0:
        print("ERROR: No years have been processed yet")
        return None

    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, Border, Side
    from io import BytesIO
        
    # Create a new workbook and select the active sheet
    wb = Workbook()
    ws = wb.active
    ws.title = "Gross Total Pay"

    # Sort years from lowest to highest
    sorted_years = sorted(years)
    
    # Sort employees alphabetically in ascending order
    sorted_employees = sorted(employees.keys())

    # Create first sheet: Gross Total Pay
    # Headers: Years (lowest to highest)
    for col_idx, year in enumerate(sorted_years, start=2):  # Start from column B
        ws.cell(row=1, column=col_idx, value=year)
    
    # Column A header: Employee Name
    ws.cell(row=1, column=1, value="Employee Name")
    
    # Populate employee data for Gross Total Pay
    row_idx = 2
    for employee_name in sorted_employees:
        ws.cell(row=row_idx, column=1, value=employee_name)
        
        # Populate gross pay data for each year
        for col_idx, year in enumerate(sorted_years, start=2):
            if year in employees[employee_name]:
                gross_pay = employees[employee_name][year][0]  # First element is gross pay
                ws.cell(row=row_idx, column=col_idx, value=gross_pay)
            else:
                ws.cell(row=row_idx, column=col_idx, value=0.00)
        
        row_idx += 1

    # Create second sheet: Employer Taxes and Contributions
    ws2 = wb.create_sheet("Employer Taxes and Contributions")
    
    # Headers: Years (lowest to highest)
    for col_idx, year in enumerate(sorted_years, start=2):  # Start from column B
        ws2.cell(row=1, column=col_idx, value=year)
    
    # Column A header: Employee Name
    ws2.cell(row=1, column=1, value="Employee Name")
    
    # Populate employee data for Employer Taxes and Contributions
    row_idx = 2
    for employee_name in sorted_employees:
        ws2.cell(row=row_idx, column=1, value=employee_name)
        
        # Populate employer taxes data for each year
        for col_idx, year in enumerate(sorted_years, start=2):
            if year in employees[employee_name]:
                employer_taxes = employees[employee_name][year][1]  # Second element is employer taxes
                ws2.cell(row=row_idx, column=col_idx, value=employer_taxes)
            else:
                ws2.cell(row=row_idx, column=col_idx, value=0.00)
        
        row_idx += 1

    # Save to BytesIO object
    databook = BytesIO()
    wb.save(databook)
    databook.seek(0)
    
    return databook
